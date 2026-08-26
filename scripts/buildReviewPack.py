#!/usr/bin/env python3
"""
Builds the advocate review pack from reviewpack.json.

One row per clause, ordered so that the clauses reaching the most documents come
first. The advocate fills three columns; importReviewSignoff.mjs reads them back
into the clause library. Nothing else in the sheet is read on import.
"""
import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
BAND_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

rows = json.load(open("reviewpack.json", encoding="utf-8"))
used = [r for r in rows if r["doc_types_reached"] > 0]
unused = [r for r in rows if r["doc_types_reached"] == 0]

# Figures quoted in the guide are computed, never typed. The previous guide still
# said "22 document types" and "61%" long after both had changed.
n_types = len({t for r in rows for t in (r["doc_types"] or "").split(", ") if t})
n_statute = sum(1 for r in used if r.get("statute_currency"))
n_questions = sum(1 for r in used if r.get("authoring_note"))
_total = sum(r["doc_types_reached"] for r in rows) or 1
_cov = lambda n: round(100 * sum(r["doc_types_reached"] for r in used[:n]) / _total)
pct10, pct20, pct50 = _cov(10), _cov(20), _cov(50)

wb = Workbook()

# ── Instructions ────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "How to use"
guide = [
    ("LegalAId — Advocate Review Pack", 16, True),
    ("", 11, False),
    ("Three columns are yours. Everything else is context.", 11, True),
    ("", 11, False),
    ("QUESTION FOR YOU  The drafter could not decide this. It names the specific", 11, False),
    ("                  judgement, the section it turns on, and anything cited but", 11, False),
    ("                  not verified. Read this column before the clause text.", 11, False),
    ("", 11, False),
    ("DECISION      Approve / Amend / Reject / Needs discussion", 11, False),
    ("REVISED TEXT  Only if you chose Amend. Paste the clause as it should read.", 11, False),
    ("              Use a blank line between numbered sub-clauses; they are", 11, False),
    ("              renumbered automatically (5.1, 5.2, 5.2.1).", 11, False),
    ("NOTE          Optional. Why, or what the engineer must change elsewhere.", 11, False),
    ("", 11, False),
    ("Sign once, at the top of the 'Reviewed clauses' tab, in the yellow cells.", 11, True),
    ("Your name goes into every clause you approved, as the reviewer of record.", 11, False),
    ("", 11, False),
    ("Where to start", 12, True),
    ("", 11, False),
    ("1. URGENT — the rows marked in the 'Question' column about the labour Codes.", 11, True),
    (f"   {n_statute} clauses were repointed after India's four labour Codes came into force on", 11, False),
    ("   21 November 2025, repealing 29 central statutes this library had been citing.", 11, False),
    ("   The SUBSTANCE moved, not just the section numbers: gratuity at one year for", 11, False),
    ("   fixed-term staff, the 50 per cent wage-definition rule, the factory day cut", 11, False),
    ("   from nine hours to eight, wages on exit within two working days. Several", 11, False),
    ("   successor sections could not be verified and are named in the Question column.", 11, False),
    ("", 11, False),
    ("2. FASTEST COVERAGE — the first rows, which are general provisions.", 11, True),
    (f"   The first 10 rows cover {pct10}% of all clause usage across the {n_types} document types.", 11, False),
    (f"   The first 20 cover {pct20}%. The first 50 cover {pct50}%.", 11, False),
    ("   Stopping after 20 or 50 is a legitimate first pass — coverage is reported", 11, False),
    ("   per document, so a partly-reviewed library is honest rather than broken.", 11, False),
    ("", 11, False),
    (f"3. THE {n_questions} ROWS WITH A QUESTION in column H.", 11, True),
    ("   These are the ones where a decision is already framed for you: the specific", 11, False),
    ("   judgement, the section it turns on, and anything cited but not verified.", 11, False),
    ("   They are quicker to decide than a clause presented as bare text.", 11, False),
    ("", 11, False),
    ("What 'reach' means", 12, True),
    (f"How many of the {n_types} document types actually render this clause in a baseline", 11, False),
    (f"draft. A clause with reach {n_types} appears in every document the product produces.", 11, False),
    ("", 11, False),
    (f"The 'Not in a baseline draft' tab lists {len(unused)} clauses that did not appear", 11, True),
    ("in the baseline sample. Most are CONDITIONAL — they are wired to a blueprint and", 11, True),
    ("appear as soon as a user answers the question that triggers them, so they still", 11, True),
    ("need review. The tab says which is which. Do not treat that tab as a delete list.", 11, True),
]
for i, (text, size, bold) in enumerate(guide, start=1):
    c = ws.cell(row=i, column=1, value=text)
    c.font = Font(name=FONT, size=size, bold=bold)
ws.column_dimensions["A"].width = 92

# ── Review sheet ────────────────────────────────────────────────────────────
rv = wb.create_sheet("Reviewed clauses")
rv["A1"] = "Reviewer name (goes into every approved clause)"
rv["A1"].font = Font(name=FONT, size=11, bold=True)
rv["C1"].fill = INPUT_FILL; rv["C1"].border = BORDER
rv["A2"] = "Review date (YYYY-MM-DD)"
rv["A2"].font = Font(name=FONT, size=11, bold=True)
rv["C2"].fill = INPUT_FILL; rv["C2"].border = BORDER
rv["A3"] = "Bar Council enrolment number (optional, recorded with the sign-off)"
rv["A3"].font = Font(name=FONT, size=11, bold=True)
rv["C3"].fill = INPUT_FILL; rv["C3"].border = BORDER

HEADERS = ["#", "Clause ID", "Title", "Reach", "Risk", "Mandatory", "Words",
           "QUESTION FOR YOU (from the drafter)", "Statutory basis", "Document types",
           "Clause text as generated",
           "DECISION", "REVISED TEXT (only if amending)", "NOTE"]
HEAD_ROW = 5
for col, name in enumerate(HEADERS, start=1):
    c = rv.cell(row=HEAD_ROW, column=col, value=name)
    c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    c.fill = HEAD_FILL
    c.alignment = Alignment(vertical="center", wrap_text=True)
    c.border = BORDER
rv.freeze_panes = "C6"

for idx, r in enumerate(used, start=1):
    row = HEAD_ROW + idx
    values = [idx, r["clause_id"], r["title"], r["doc_types_reached"],
              r["risk_level"], "Yes" if r["mandatory"] else "", r["rendered_words"],
              r.get("authoring_note", ""),
              r["citations"], r["doc_types"], r["text"], "", "", ""]
    for col, v in enumerate(values, start=1):
        c = rv.cell(row=row, column=col, value=v)
        c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(vertical="top", wrap_text=col in (3, 8, 9, 10, 11, 13, 14))
        c.border = BORDER
        # The drafter's question is the column the advocate should read first.
        if col == 8 and v:
            c.font = Font(name=FONT, size=10, color="7F3F00")
        if col in (12, 13, 14):
            c.fill = INPUT_FILL
        elif idx % 2 == 0:
            c.fill = BAND_FILL

dv = DataValidation(type="list", formula1='"Approve,Amend,Reject,Needs discussion"', allow_blank=True)
rv.add_data_validation(dv)
dv.add(f"L{HEAD_ROW+1}:L{HEAD_ROW+len(used)}")

WIDTHS = {1:5, 2:34, 3:26, 4:7, 5:9, 6:10, 7:8, 8:58, 9:34, 10:32,
          11:70, 12:18, 13:58, 14:32}
for col, w in WIDTHS.items():
    rv.column_dimensions[get_column_letter(col)].width = w

# ── Unused clauses ──────────────────────────────────────────────────────────
uw = wb.create_sheet("Not in a baseline draft")
uw["A1"] = (
    f"{len(unused)} clauses did not appear in the baseline sample of "
    f"{len(set(t for r in rows for t in (r['doc_types'] or '').split(', ') if t))} document types. "
    "Check the 'Why' column: a conditional clause is live and still needs review."
)
uw["A1"].font = Font(name=FONT, size=12, bold=True)
UH = ["Clause ID", "Title", "Domain", "Risk", "Why it did not appear", "Statutory basis",
      "KEEP / DELETE", "NOTE"]
for col, name in enumerate(UH, start=1):
    c = uw.cell(row=3, column=col, value=name)
    c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    c.fill = HEAD_FILL; c.border = BORDER
for i, r in enumerate(unused, start=1):
    values = [r["clause_id"], r["title"], r["domain"], r["risk_level"],
              r.get("reach_status", ""), r["citations"], "", ""]
    for col, v in enumerate(values, start=1):
        c = uw.cell(row=3 + i, column=col, value=v)
        c.font = Font(name=FONT, size=10); c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=(col in (2, 5, 6, 8)))
        if col in (7, 8): c.fill = INPUT_FILL
for col, w in {1:34, 2:28, 3:13, 4:9, 5:46, 6:34, 7:16, 8:30}.items():
    uw.column_dimensions[get_column_letter(col)].width = w

out = sys.argv[1] if len(sys.argv) > 1 else "LegalAId_Advocate_Review_Pack.xlsx"
wb.save(out)
print(f"wrote {out}: {len(used)} clauses to review, {len(unused)} unused")
