#!/usr/bin/env python3
"""
Reads a completed advocate review pack back into the clause library.

Writes reviewed_by / reviewed_on / review_status onto each approved clause, and
applies revised text where the advocate amended one. A clause counts as reviewed
only when it carries a real reviewer name AND its status is no longer
"draft-needs-legal-review" -- see shared/clauseProvenance.js.

Usage:  python3 scripts/importReviewSignoff.py <pack.xlsx> [--apply]
Without --apply it reports what it would do and changes nothing.
"""
import json, sys, os
from openpyxl import load_workbook

PLACEHOLDERS = {"", "pending", "tbd", "todo", "none", "n/a", "unknown"}

def main():
    if len(sys.argv) < 2:
        sys.exit("usage: importReviewSignoff.py <pack.xlsx> [--apply]")
    pack, apply_changes = sys.argv[1], "--apply" in sys.argv

    wb = load_workbook(pack)
    rv = wb["Reviewed clauses"]
    reviewer = str(rv["C1"].value or "").strip()
    review_date = str(rv["C2"].value or "").strip()[:10]
    enrolment = str(rv["C3"].value or "").strip()

    if reviewer.lower() in PLACEHOLDERS:
        sys.exit("No reviewer name in cell C1. That name is the signature — "
                 "without it nothing counts as reviewed.")
    if not review_date:
        sys.exit("No review date in cell C2 (expected YYYY-MM-DD).")

    index = {c["clause_id"]: c["file"] for c in json.load(open("reviewpack.json", encoding="utf-8"))}

    # Locate the columns by their HEADER TEXT rather than by position. They were
    # read by fixed index until a column was inserted ahead of them, at which
    # point the importer silently read the wrong three columns and reported every
    # clause as skipped. A header lookup cannot fail that way.
    header_row = 5
    headers = {}
    for cell in rv[header_row]:
        if cell.value:
            headers[str(cell.value).strip().upper()] = cell.column - 1

    def column_for(*candidates):
        for want in candidates:
            for name, idx in headers.items():
                if name.startswith(want):
                    return idx
        sys.exit(
            f"Could not find the '{candidates[0]}' column in row {header_row} of the "
            f"'Reviewed clauses' tab. Found: {sorted(headers)}. The sheet was probably "
            "produced by a different version of scripts/buildReviewPack.py."
        )

    COL_ID = column_for("CLAUSE ID")
    COL_DECISION = column_for("DECISION")
    COL_REVISED = column_for("REVISED TEXT")
    COL_NOTE = column_for("NOTE")

    approved, amended, rejected, discuss, skipped, missing = [], [], [], [], [], []
    edits = {}

    for row in rv.iter_rows(min_row=6, values_only=True):
        if not row or len(row) <= COL_ID or not row[COL_ID]:
            continue
        cell = lambda i: row[i] if i < len(row) else None
        clause_id = str(row[COL_ID]).strip()
        decision = str(cell(COL_DECISION) or "").strip().lower()
        revised = cell(COL_REVISED) or ""
        note = str(cell(COL_NOTE) or "").strip()

        if not decision:
            skipped.append(clause_id); continue
        if decision.startswith("reject"):
            rejected.append((clause_id, note)); continue
        if decision.startswith("needs"):
            discuss.append((clause_id, note)); continue
        if not decision.startswith(("approve", "amend")):
            skipped.append(clause_id); continue

        path = index.get(clause_id)
        if not path or not os.path.exists(path):
            missing.append(clause_id); continue

        entry = edits.setdefault(path, [])
        entry.append({
            "clause_id": clause_id,
            "text": str(revised).strip() if decision.startswith("amend") else None,
            "note": note,
        })
        (amended if decision.startswith("amend") else approved).append(clause_id)

    print(f"reviewer      : {reviewer}{f' (enrolment {enrolment})' if enrolment else ''}")
    print(f"review date   : {review_date}")
    print(f"approved      : {len(approved)}")
    print(f"amended       : {len(amended)}")
    print(f"rejected      : {len(rejected)}")
    print(f"needs discuss : {len(discuss)}")
    print(f"no decision   : {len(skipped)}")
    if missing:
        print(f"NOT FOUND     : {len(missing)} -> {', '.join(missing[:8])}")
    for label, items in (("REJECTED", rejected), ("NEEDS DISCUSSION", discuss)):
        if items:
            print(f"\n{label}:")
            for cid, note in items:
                print(f"  {cid}: {note or '(no note)'}")

    if not apply_changes:
        print("\nDry run. Re-run with --apply to write these into the clause library.")
        return

    written = 0
    for path, items in edits.items():
        raw = json.load(open(path, encoding="utf-8"))
        listed = raw if isinstance(raw, list) else [raw]
        by_id = {c.get("clause_id"): c for c in listed if isinstance(c, dict)}
        for item in items:
            clause = by_id.get(item["clause_id"])
            if clause is None:
                continue
            clause["reviewed_by"] = reviewer
            clause["reviewed_on"] = review_date
            clause["review_status"] = "approved"
            if enrolment:
                clause["reviewer_enrolment"] = enrolment
            if item["text"]:
                clause["text"] = item["text"]
            if item["note"]:
                clause["review_note"] = item["note"]
            written += 1
        json.dump(raw, open(path, "w", encoding="utf-8"), indent=2, ensure_ascii=False)
        open(path, "a", encoding="utf-8").write("\n")

    print(f"\nwrote sign-off onto {written} clauses across {len(edits)} files")
    print("Now run: node tests/clauseProvenance.test.mjs   (the reviewed count should rise)")

main()
